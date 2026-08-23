# -*- coding: utf-8 -*-
"""
Transforme un export de métré (.json produit par l'app) en fiche OIKEN (.xlsx).

Le fichier est une COPIE du modèle `data/sources/modele-fiche-oiken.xlsx` :
logo, styles, largeurs de colonnes, zone d'impression et onglets de référence
(« Codes CI », « Aide-mémoire CAN ») sont conservés à l'identique. Seules les
cellules de la fiche sont remplies.

Usage :
    python scripts/metre-vers-oiken.py <export.json> [dossier_de_sortie]

Sans dossier de sortie, le fichier est écrit sur le Bureau.
Le nom du fichier reprend le nom de l'affaire ; un fichier existant n'est jamais
écrasé (un suffixe -2, -3… est ajouté).
"""
import json
import os
import re
import shutil
import sys
import unicodedata
from copy import copy

import openpyxl

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODELE = os.path.join(RACINE, 'data', 'sources', 'modele-fiche-oiken.xlsx')

PREMIERE_LIGNE = 10          # première ligne du tableau, sous les en-têtes
DERNIERE_LIGNE_MODELE = 49   # fin de la zone d'impression d'origine
LIGNE_MODELE_DONNEE = 14     # ligne dont on clone le style pour les ajouts

COL_OBJET, COL_CI, COL_CAN, COL_QTE, COL_CHF = 'A', 'E', 'F', 'H', 'J'
FUSIONS = [('A', 'D'), ('F', 'G'), ('H', 'I'), ('J', 'K')]


def nom_de_fichier(nom):
    """Nom sûr sur tous les systèmes, accents retirés."""
    sans_accents = ''.join(
        c for c in unicodedata.normalize('NFD', nom) if unicodedata.category(c) != 'Mn'
    )
    propre = re.sub(r'[^A-Za-z0-9 _-]', '', sans_accents).strip()
    return re.sub(r'\s+', ' ', propre) or 'Sans nom'


def chemin_libre(dossier, base):
    chemin = os.path.join(dossier, base + '.xlsx')
    n = 2
    while os.path.exists(chemin):
        chemin = os.path.join(dossier, '%s-%d.xlsx' % (base, n))
        n += 1
    return chemin


def vider_zone(ws, debut, fin):
    """
    Efface les valeurs du tableau sans toucher aux styles ni aux fusions.
    Dans une plage fusionnée, seule la cellule en haut à gauche est inscriptible :
    les autres sont des MergedCell en lecture seule, qu'on saute.
    """
    for ligne in range(debut, fin + 1):
        for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
            cellule = ws['%s%d' % (col, ligne)]
            if isinstance(cellule, openpyxl.cell.cell.Cell):
                cellule.value = None


def cloner_style_ligne(ws, source, cible):
    for col in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
        src = ws['%s%d' % (col, source)]
        dst = ws['%s%d' % (col, cible)]
        dst._style = copy(src._style)
    ws.row_dimensions[cible].height = ws.row_dimensions[source].height


def fusionner(ws, ligne):
    existantes = {str(m) for m in ws.merged_cells.ranges}
    for gauche, droite in FUSIONS:
        plage = '%s%d:%s%d' % (gauche, ligne, droite, ligne)
        if plage not in existantes:
            try:
                ws.merge_cells(plage)
            except ValueError:
                pass


def remplir(donnees, ws):
    affaire = donnees['affaire']

    # Le modèle est livré avec un exemple rempli : sans ça, les lignes d'exemple
    # qui dépassent le nouveau métré restent dans la fiche.
    vider_zone(ws, PREMIERE_LIGNE, DERNIERE_LIGNE_MODELE)

    # En-tête : le modèle attend le nom du client à droite du libellé « Client ».
    ws['B5'] = affaire['nom']
    entete = [affaire.get('numero_affaire'), affaire.get('localite')]
    ws['B6'] = ' · '.join(x for x in entete if x)

    ligne = PREMIERE_LIGNE
    derniere_stylee = DERNIERE_LIGNE_MODELE

    for bloc in donnees['locaux']:
        if ligne > derniere_stylee:
            cloner_style_ligne(ws, LIGNE_MODELE_DONNEE, ligne)
            derniere_stylee = ligne
        fusionner(ws, ligne)
        ws['%s%d' % (COL_OBJET, ligne)] = bloc['nom'].upper()
        ligne += 1

        for l in bloc['lignes']:
            if ligne > derniere_stylee:
                cloner_style_ligne(ws, LIGNE_MODELE_DONNEE, ligne)
                derniere_stylee = ligne
            fusionner(ws, ligne)
            objet = '    ' + l['objet']
            if l['unite'] != 'pce':
                objet += '  [%s]' % l['unite']
            ws['%s%d' % (COL_OBJET, ligne)] = objet
            ws['%s%d' % (COL_CI, ligne)] = l['ci'] or None
            ws['%s%d' % (COL_CAN, ligne)] = l['no_can'] or None
            ws['%s%d' % (COL_QTE, ligne)] = l['quantite']
            ligne += 1

        ligne += 1  # une ligne vide entre deux locaux, comme sur le modèle

    fin = max(ligne, DERNIERE_LIGNE_MODELE)
    ws.print_area = 'A1:K%d' % fin
    return ligne - PREMIERE_LIGNE


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    source = sys.argv[1]
    dossier = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.expanduser('~'), 'OneDrive', 'Bureau'
    )
    if not os.path.isdir(dossier):
        dossier = os.path.expanduser('~')

    donnees = json.load(open(source, encoding='utf-8'))
    base = 'Metre - ' + nom_de_fichier(donnees['affaire']['nom'])
    sortie = chemin_libre(dossier, base)

    # On copie d'abord, puis on modifie la copie : le modèle reste intact.
    shutil.copyfile(MODELE, sortie)
    wb = openpyxl.load_workbook(sortie)
    lignes = remplir(donnees, wb['Fiche'])
    wb.save(sortie)

    nb = sum(len(b['lignes']) for b in donnees['locaux'])
    print('Fiche generee : %s' % sortie)
    print('  affaire  : %s' % donnees['affaire']['nom'])
    print('  locaux   : %d' % len(donnees['locaux']))
    print('  lignes   : %d  (%d lignes de tableau utilisees)' % (nb, lignes))
    sans_ci = sum(1 for b in donnees['locaux'] for l in b['lignes'] if not l['ci'])
    if sans_ci:
        print('  ATTENTION : %d ligne(s) sans code CI' % sans_ci)


if __name__ == '__main__':
    main()
